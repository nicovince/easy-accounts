"""Manage Account Spreadsheet."""

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.comments import Comment


class AccountSpreadsheet:
    """Class to manage an account spreadsheet."""

    def __init__(self, spreadsheet_path: str):
        self.path = spreadsheet_path
        self.wb = openpyxl.load_workbook(self.path)
        self.category_column = "A"
        self.month_row = 1
        self.user_row = 2
        self._active_sheet = None

    @property
    def active_sheet(self):
        return self._active_sheet

    @active_sheet.setter
    def active_sheet(self, value):
        assert value in self.wb.sheetnames
        self._active_sheet = value

    def save(self):
        """Save file to disk."""
        self.wb.save(self.path)

    def get_sheet(self):
        """Get the requested sheet."""
        if self.active_sheet is None:
            return self.wb.active
        return self.wb[self.active_sheet]

    def get_cell_category(self, category: str) -> Cell:
        """Get the cell for a matching category."""
        ws = self.get_sheet()
        categories = ws[self.category_column]
        for cell in categories:
            if cell.value == category:
                return cell
        assert False, f"{category} not found in {self.path}"

    def get_cell_month(self, month: str) -> Cell:
        """Return the cell for a matching month."""
        ws = self.get_sheet()
        months = ws[self.month_row]
        for cell in months:
            if cell.value == month:
                return cell
        assert False, f"{month} not found in {self.path}"

    def get_next_cell(self, cell: Cell) -> Cell:
        """Get cell in the next column."""
        ws = self.get_sheet()
        return ws.cell(row=cell.row, column=cell.column + 1)

    def get_next_month_cell(self, month_cell: Cell) -> Cell:
        """Get the next month cell of the provided month cell"""
        assert month_cell.row == self.month_row
        nxt_cell = self.get_next_cell(month_cell)
        while type(self.get_next_cell(nxt_cell)) is openpyxl.cell.cell.MergedCell:
            nxt_cell = self.get_next_cell(nxt_cell)
        return self.get_next_cell(nxt_cell)

    def get_cell(self, month: str, category: str, user=None) -> Cell:
        """Get cell for matching month and category."""
        month_cell = self.get_cell_month(month)
        column = month_cell.col_idx
        category_cell = self.get_cell_category(category)
        ws = self.get_sheet()
        if user is not None:
            nxt_month = self.get_next_month_cell(month_cell)
            user_range_start = f"{month_cell.column_letter}{self.user_row}"
            user_range_end = f"{nxt_month.column_letter}{self.user_row}"
            month_users_range = ws[user_range_start:user_range_end]
            for row in month_users_range:
                for user_cell in row:
                    if user_cell.value == user:
                        column = user_cell.col_idx
                        break

        return ws.cell(row=category_cell.row, column=column)

    def add_entry(
        self,
        month: str,
        category: str,
        amount: float | int | list[float] | list[int],
        comment: str | None = None,
        user: str | None = None,
    ) -> None:
        """Add an account entry."""
        cell = self.get_cell(month, category, user)

        # Convert amount to list if it's a single float/int
        if isinstance(amount, (float, int)):
            amounts = [amount]
        else:
            amounts = amount

        # Build the amount string: "a + b + c + d"
        amount_str = " + ".join(str(a) for a in amounts)

        if cell.value is None:
            cell.value = f"={amount_str}"
        else:
            cell.value = f"{cell.value} + {amount_str}"
        if comment is not None:
            if cell.comment is not None:
                cell.comment.text += f"\n{comment}"
            else:
                cell.comment = Comment(comment, "easy-account")

    def get_spreadsheet_months(self) -> list[str]:
        """Get list of months from spreadsheet.

        Returns:
            List of month names found in the spreadsheet.
        """
        ws = self.get_sheet()
        months: list[str] = []
        for cell in ws[self.month_row]:
            if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                months.append(str(cell.value))
        return months

    def get_spreadsheet_categories(self) -> list[str]:
        """Get list of categories from spreadsheet.

        Returns:
            List of category names found in the spreadsheet.
        """
        ws = self.get_sheet()
        categories: list[str] = []
        for cell in ws[self.category_column]:
            if cell.value:
                categories.append(str(cell.value))
        return categories

    def get_spreadsheet_users(self) -> list[str]:
        """Get list of users from spreadsheet.

        Returns:
            List of user names found in the spreadsheet, or empty list if no users.
        """
        ws = self.get_sheet()
        users: set[str] = set()
        month_columns: set[int] = set()

        for cell in ws[self.month_row]:
            if isinstance(cell, openpyxl.cell.cell.MergedCell):
                continue
            if cell.value:
                month_columns.add(cell.column)
                for merged_range in ws.merged_cells.ranges:
                    if (
                        merged_range.min_row <= 1 <= merged_range.max_row
                        and merged_range.min_col <= cell.column <= merged_range.max_col
                    ):
                        for col in range(merged_range.min_col, merged_range.max_col + 1):
                            month_columns.add(col)
                        break

        for row in ws.iter_rows(min_row=self.user_row, max_row=self.user_row):
            for cell in row:
                if (
                    cell.column in month_columns
                    and cell.value
                    and not isinstance(cell, openpyxl.cell.cell.MergedCell)
                ):
                    users.add(str(cell.value))
        return list(users)

    def is_multiuser(self) -> bool:
        """Check if spreadsheet is multi-user based on merged cells in month row.

        Returns:
            True if spreadsheet has merged cells in the month row (multi-user).
        """
        ws = self.get_sheet()
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row == 1 and merged_range.max_row == 1:
                return True
        return False
