import sys
import shutil
import os
import hashlib
from pathlib import Path
from openpyxl import Workbook

import pytest

project_root = Path(__file__).parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

os.chdir(project_root)


def get_months():
    months = [
        "janvier",
        "fevrier",
        "mars",
        "avril",
        "mai",
        "juin",
        "aout",
        "septembre",
        "octobre",
        "novembre",
        "decembre",
    ]
    return months


def get_categories():
    return ["out-foo", "out-bar", "all-out", "in-foo", "in-bar", "all-in", "balance"]


def get_users():
    return ["alice", "bob", "shared"]


def fill_monouser_sheet(ws):
    months = get_months()
    col_month_offset = 2
    row_category_offset = 2
    categories = get_categories()
    all_out_row = row_category_offset + categories.index("all-out")
    all_in_row = row_category_offset + categories.index("all-in")
    balance_row = row_category_offset + categories.index("balance")
    for idx, month in enumerate(months):
        c = ws.cell(row=1, column=(col_month_offset + idx), value=month)
        cur_col = c.column_letter
        ws.cell(
            row=all_out_row,
            column=(col_month_offset + idx),
            value=f"=SUM({cur_col}{row_category_offset}:{cur_col}{all_out_row - 1}",
        )
        ws.cell(
            row=all_in_row,
            column=(col_month_offset + idx),
            value=f"=SUM({cur_col}{all_out_row + 1}:{cur_col}{all_in_row - 1}",
        )
        ws.cell(
            row=balance_row,
            column=(col_month_offset + idx),
            value=f"={cur_col}{all_in_row} - {cur_col}{all_out_row}",
        )

    for idx, category in enumerate(categories):
        ws.cell(row=(row_category_offset + idx), column=1, value=category)

    # janvier / out-bar
    ws["B3"] = "=1234"
    # janvier / in-bar
    ws["B6"] = "=234"
    # fevrier / in-foo
    ws["C5"] = "=111"


def fill_multiuser_sheet(ws):
    users = get_users()
    categories = get_categories()
    col_month_offset = 2
    row_category_offset = 3
    all_out_row = row_category_offset + categories.index("all-out")
    all_in_row = row_category_offset + categories.index("all-in")
    balance_row = row_category_offset + categories.index("balance")
    months = get_months()
    for idx, month in enumerate(months):
        month_start_col = col_month_offset + idx * len(users)
        month_end_col = col_month_offset + (idx + 1) * len(users) - 1
        ws.cell(row=1, column=month_start_col, value=month)
        ws.merge_cells(
            start_row=1, start_column=month_start_col, end_row=1, end_column=month_end_col
        )
        for user_idx, user in enumerate(users):
            c = ws.cell(row=2, column=(col_month_offset + idx * len(users) + user_idx), value=user)
            cur_col = c.column_letter
            ws.cell(
                row=all_out_row,
                column=c.col_idx,
                value=f"=SUM({cur_col}{row_category_offset}:{cur_col}{all_out_row - 1}",
            )
            ws.cell(
                row=all_in_row,
                column=c.col_idx,
                value=f"=SUM({cur_col}{all_out_row + 1}:{cur_col}{all_in_row - 1}",
            )
            ws.cell(
                row=balance_row,
                column=c.col_idx,
                value=f"={cur_col}{all_in_row} - {cur_col}{all_out_row}",
            )

    for idx, category in enumerate(categories):
        ws.cell(row=(row_category_offset + idx), column=1, value=category)
    # Janvier / Alice / out-bar
    ws["B4"] = "=100"
    # Janvier / Bob / out-bar
    ws["C4"] = "=200"
    # Janvier / Shared / out-bar
    ws["D4"] = "=300"
    # Janvier / Alice / in-bar
    ws["B7"] = "=150"
    # Janvier / Bob / in-bar
    ws["C7"] = "=220"
    # Janvier / Shared / in-bar
    ws["D7"] = "=360"

    # Decembre / Alice / out-foo
    ws["AF3"] = "=12*100"
    # Decembre / Bob / out-foo
    ws["AG3"] = "=12*200"
    # Decembre / Shared / out-foo
    ws["AH3"] = "=12*300"
    # Decembre / Alice / in-foo
    ws["AF6"] = "=12*101"
    # Decembre / Bob / in-foo
    ws["AG6"] = "=12*202"
    # Decembre / Shared / in-foo
    ws["AH6"] = "=12*303"


@pytest.fixture
def tmp_path_cwd(tmp_path):
    """Set current working directory to tmp_path and restore on teardown"""
    original_cwd = os.getcwd()
    os.chdir(tmp_path)
    yield tmp_path
    os.chdir(original_cwd)


def create_spreadsheet_template(filename: str) -> None:
    """Create a spreadsheet with both mono-user and multi-user sheets."""
    wb = Workbook()
    wb.remove(wb.active)
    fill_monouser_sheet(wb.create_sheet("mono user"))
    fill_multiuser_sheet(wb.create_sheet("multi users"))
    wb.save(filename)


@pytest.fixture(scope="session")
def spreadsheet_template(tmp_path_factory):
    """Fixture to create spreadsheet template."""
    path = tmp_path_factory.mktemp("template") / "template_spreadsheet.xlsx"
    create_spreadsheet_template(path)
    return str(path)


@pytest.fixture
def spreadsheet_unmodified(tmp_path, spreadsheet_template):
    """Spreadhseet which must not be modified by the test."""
    path = tmp_path / "unmodified_spreadsheet.xlsx"
    shutil.copy(spreadsheet_template, path)
    with open(str(path), "rb") as f:
        orig_digest = hashlib.sha256(f.read())
    yield str(path)
    with open(str(path), "rb") as f:
        digest = hashlib.sha256(f.read())
    assert orig_digest.hexdigest() == digest.hexdigest()


@pytest.fixture
def spreadsheet(tmp_path, spreadsheet_template):
    path = tmp_path / "test_spreadsheet.xlsx"
    shutil.copy(spreadsheet_template, path)
    return str(path)


@pytest.fixture
def monouser_account(spreadsheet):
    from easy_account.account import AccountSpreadsheet

    acc = AccountSpreadsheet(spreadsheet)
    acc.active_sheet = "mono user"
    return spreadsheet


@pytest.fixture
def multiuser_account_fixture(spreadsheet):
    return spreadsheet


@pytest.fixture
def fresh_monouser_account(tmp_path):
    wb = Workbook()
    ws = wb.active
    fill_monouser_sheet(ws)
    path = tmp_path / "fresh_monouser_account.xlsx"
    wb.save(path)
    return str(path)


@pytest.fixture
def fresh_multiuser_account(tmp_path):
    wb = Workbook()
    ws = wb.active
    fill_multiuser_sheet(ws)
    path = tmp_path / "fresh_multiuser_account.xlsx"
    wb.save(path)
    return str(path)
