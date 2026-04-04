import pytest
from openpyxl import Workbook
from easy_account.spreadsheet import Spreadsheet
from easy_account.spreadsheet import CellRange


@pytest.fixture(scope="session")
def sample_spreadsheet(tmp_path_factory):
    path = f"{tmp_path_factory.mktemp('sample_spreadsheet')}/sample_spreadsheet.xlsx"
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)
    wb.create_sheet("Sheet1")
    wb.create_sheet("Sheet2")
    wb.save(path)
    return Spreadsheet(path)


class TestCellRange:
    def test_cell_range_sheet_name(self):
        cr = CellRange("A1")
        assert cr.get_parent_sheet_name() is None
        cr = CellRange("Sheet1!A1")
        assert cr.get_parent_sheet_name() == "Sheet1"

    def test_cell_range_single_cell(self):
        assert CellRange("A1").is_single_cell()
        assert CellRange("Sheet1!A1").is_single_cell()
        assert not CellRange("A1:B2").is_single_cell()
        assert not CellRange("Sheet1!A1:B2").is_single_cell()

    def test_cell_range_get_start_pos(self):
        assert CellRange("A1").get_start_pos() == "A1"
        assert CellRange("Sheet1!A1").get_start_pos() == "A1"
        assert CellRange("A1:B2").get_start_pos() == "A1"
        assert CellRange("Sheet1!A1:B2").get_start_pos() == "A1"

    def test_cell_range_get_end_pos(self):
        assert CellRange("A1:B2").get_end_pos() == "B2"
        assert CellRange("Sheet1!A1:B2").get_end_pos() == "B2"


class TestSpreadsheetHelpers:
    """Test helpers from Spreadsheet"""

    def test_spreadsheet_get_sheet_by_name(self, sample_spreadsheet):
        ws = sample_spreadsheet.get_sheet("Sheet1")
        assert ws.title == "Sheet1"

    def test_spreadsheet_get_cell_value(self, sample_spreadsheet):
        ws = sample_spreadsheet.get_sheet("Sheet1")
        ws["A1"] = 5
        val = sample_spreadsheet.get_cell_value("Sheet1", "A", 1)
        assert val == 5


class TestSpreadsheetEvaluate:
    """Test Spreadsheet Evaluation"""

    def test_spreadsheet_evaluate_no_formula_int(self, sample_spreadsheet):
        ws = sample_spreadsheet.get_sheet("Sheet1")
        ws["A1"] = "5"
        val = sample_spreadsheet.evaluate(ws["A1"])
        assert val == 5

    def test_spreadsheet_evaluate_no_formula_float(self, sample_spreadsheet):
        ws = sample_spreadsheet.get_sheet("Sheet1")
        ws["A1"] = "5.2"
        val = sample_spreadsheet.evaluate(ws["A1"])
        assert val == 5.2

    def test_spreadsheet_evaluate_formula_int(self, sample_spreadsheet):
        ws = sample_spreadsheet.get_sheet("Sheet1")
        ws["A1"] = "=3"
        val = sample_spreadsheet.evaluate(ws["A1"])
        assert val == 3

    def test_spreadsheet_evaluate_formula_float(self, sample_spreadsheet):
        ws = sample_spreadsheet.get_sheet("Sheet1")
        ws["A1"] = "=3.14"
        val = sample_spreadsheet.evaluate(ws["A1"])
        assert val == 3.14

    def test_spreadsheet_evaluate_formula_add_int(self, sample_spreadsheet):
        ws = sample_spreadsheet.get_sheet("Sheet1")
        ws["A1"] = "=5+3"
        val = sample_spreadsheet.evaluate(ws["A1"])
        assert val == 8

    def test_spreadsheet_evaluate_formula_add_float(self, sample_spreadsheet):
        ws = sample_spreadsheet.get_sheet("Sheet1")
        ws["A1"] = "=0.1 + 3.4"
        val = sample_spreadsheet.evaluate(ws["A1"])
        assert val == 3.5

    def test_spreadsheet_evaluate_cell_ref(self, sample_spreadsheet):
        ws = sample_spreadsheet.get_sheet("Sheet1")
        ws["A1"] = "5"
        ws["A2"] = "=A1"
        val = sample_spreadsheet.evaluate(ws["A2"])
        assert val == 5

    def test_spreadsheet_evaluate_cell_ref_other_sheet(self, sample_spreadsheet):
        ws = sample_spreadsheet.get_sheet("Sheet1")
        ws["A1"] = "5"
        ws2 = sample_spreadsheet.get_sheet("Sheet2")
        ws2["A1"] = "=3 + Sheet1!A1"
        val = sample_spreadsheet.evaluate(ws2["A1"])
        assert val == 8

    def test_spreadsheet_evaluate_sum_range(self, sample_spreadsheet):
        ws = sample_spreadsheet.get_sheet("Sheet1")
        ws["A1"] = "=5"
        ws["A2"] = "=A1"
        ws["A3"] = "=SUM(A1:A2)"
        val = sample_spreadsheet.evaluate(ws["A3"])
        assert val == 10

    def test_spreadsheet_evaluate_multiple_sum(self, sample_spreadsheet):
        ws = sample_spreadsheet.get_sheet("Sheet1")
        ws["A1"] = "=5"
        ws["A2"] = "=A1"
        ws["A3"] = "=SUM(A1:A2)"
        ws["B1"] = "=4"
        ws["B2"] = "=B1"
        ws["B3"] = "=SUM(B1:B2)"
        ws["C1"] = "=SUM(A1:A2) + SUM(B1:B2)"
        val = sample_spreadsheet.evaluate(ws["C1"])
        assert val == 18
        ws["C2"] = "=SUM(A1:B2)"
        val = sample_spreadsheet.evaluate(ws["C2"])
        assert val == 18
        ws["C3"] = "=SUM(A1:A2) - SUM(B1:B2)"
        val = sample_spreadsheet.evaluate(ws["C3"])
        assert val == 2
