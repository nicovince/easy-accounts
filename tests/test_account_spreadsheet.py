from easy_account.account import AccountSpreadsheet
from openpyxl import Workbook
import conftest
import pytest


@pytest.fixture(scope="session")
def multiuser_account(tmp_path_factory):
    wb = Workbook()
    ws = wb.active
    conftest.fill_multiuser_sheet(ws)
    path = f"{tmp_path_factory.mktemp('multiuser')}/multi_user_account.xlsx"
    wb.save(path)
    return AccountSpreadsheet(path)


@pytest.fixture(scope="session")
def multisheet(tmp_path_factory):
    wb = Workbook()
    conftest.fill_multiuser_sheet(wb.create_sheet("multi users"))
    conftest.fill_monouser_sheet(wb.create_sheet("mono user"))
    path = f"{tmp_path_factory.mktemp('multisheet')}/multi_sheet_account.xlsx"
    wb.save(path)
    return AccountSpreadsheet(path)


@pytest.fixture
def dummy_account(monouser_account):
    return AccountSpreadsheet(monouser_account)


def test_account_constructor():
    AccountSpreadsheet("tests/dummy_account.xlsx")


def test_account_get_row_from_category(dummy_account):
    cell = dummy_account.get_cell_category("foo")
    assert cell.row == 2


def test_account_invalid_category(dummy_account):
    with pytest.raises(AssertionError):
        dummy_account.get_cell_category("dsmfkj")


def test_account_get_cell_month(dummy_account):
    cell = dummy_account.get_cell_month("janvier")
    assert cell.column_letter == "B"


def test_account_get_cell_month_invalid(dummy_account):
    with pytest.raises(AssertionError):
        dummy_account.get_cell_month("qmlfkj")


def test_account_get_cell_month_category(dummy_account):
    c = dummy_account.get_cell(month="janvier", category="foo")
    assert c.coordinate == "B2"


def test_account_get_cell_month_category_invalid_month(dummy_account):
    with pytest.raises(AssertionError):
        dummy_account.get_cell(month="pwet", category="foo")


def test_account_get_cell_month_category_invalid_category(dummy_account):
    with pytest.raises(AssertionError):
        dummy_account.get_cell(month="janvier", category="pwet")


def test_account_multi_get_month(multiuser_account):
    c = multiuser_account.get_cell_month("janvier")
    assert c.column_letter == "B"
    c = multiuser_account.get_cell_month("fevrier")
    assert c.column_letter == "E"
    c = multiuser_account.get_cell_month("decembre")
    assert c.column_letter == "AF"


def test_account_multi_next_month(multiuser_account):
    c = multiuser_account.get_cell_month("janvier")
    cn = multiuser_account.get_next_month_cell(c)
    assert cn.column_letter == "E"
    c = multiuser_account.get_cell_month("decembre")
    cn = multiuser_account.get_next_month_cell(c)
    assert cn.column_letter == "AI"


def test_account_multi_get_cell_month_category_user(multiuser_account):
    c = multiuser_account.get_cell(month="janvier", category="foo", user="alice")
    assert c.coordinate == "B3"
    c = multiuser_account.get_cell(month="janvier", category="foo", user="bob")
    assert c.coordinate == "C3"
    c = multiuser_account.get_cell(month="janvier", category="foo", user="shared")
    assert c.coordinate == "D3"

    c = multiuser_account.get_cell(month="decembre", category="foo", user="alice")
    assert c.coordinate == "AF3"
    c = multiuser_account.get_cell(month="decembre", category="foo", user="bob")
    assert c.coordinate == "AG3"
    c = multiuser_account.get_cell(month="decembre", category="foo", user="shared")
    assert c.coordinate == "AH3"


def test_account_multisheet_valid_sheet(multisheet):
    multisheet.active_sheet = "multi users"
    c = multisheet.get_cell(month="janvier", category="foo", user="alice")
    assert c.coordinate == "B3"


def test_account_multisheet_invalid_sheet(multisheet):
    with pytest.raises(AssertionError):
        multisheet.active_sheet = "invalid"


def test_account_add_entry(multisheet):
    multisheet.active_sheet = "multi users"
    c = multisheet.get_cell(month="janvier", category="foo")
    multisheet.add_entry("janvier", "foo", 3.14, "pi")
    assert c.value == "=3.14"
    assert c.comment.text == "pi"
    multisheet.add_entry("janvier", "foo", 3.14, "pi")
    assert c.value == "=3.14 + 3.14"
    assert c.comment.text == "pi\npi"
    c = multisheet.get_cell(month="fevrier", category="foo")
    multisheet.add_entry("fevrier", "foo", 3615)
    assert c.value == "=3615"
    assert c.comment is None


def test_account_add_entry_with_list_empty_cell(multisheet):
    """Test adding a list of floats to an empty cell."""
    multisheet.active_sheet = "mono user"
    c = multisheet.get_cell(month="janvier", category="foo")
    multisheet.add_entry("janvier", "foo", [1.0, 2.0, 3.0])
    assert c.value == "=1.0 + 2.0 + 3.0"
    assert c.comment is None


def test_account_add_entry_with_list_existing_cell(multisheet):
    """Test adding a list of floats to an existing cell."""
    multisheet.active_sheet = "mono user"
    c = multisheet.get_cell(month="mars", category="foo")
    multisheet.add_entry("mars", "foo", [1.0, 2.0])
    assert c.value == "=1.0 + 2.0"
    multisheet.add_entry("mars", "foo", [3.0, 4.0])
    assert c.value == "=1.0 + 2.0 + 3.0 + 4.0"


def test_account_add_entry_with_list_and_comment(multisheet):
    """Test adding a list of floats with a comment."""
    multisheet.active_sheet = "mono user"
    c = multisheet.get_cell(month="avril", category="foo")
    multisheet.add_entry("avril", "foo", [5.5, 4.5], "mixed")
    assert c.value == "=5.5 + 4.5"
    assert c.comment.text == "mixed"


def test_account_add_entry_with_single_float_in_list(multisheet):
    """Test adding a single float in a list."""
    multisheet.active_sheet = "mono user"
    c = multisheet.get_cell(month="mai", category="foo")
    multisheet.add_entry("mai", "foo", [10.0])
    assert c.value == "=10.0"


class TestAccountSpreadsheetHelpers:
    """Tests for AccountSpreadsheet helper methods."""

    def test_get_spreadsheet_months_monouser(self, dummy_account):
        """Test getting months from monouser spreadsheet."""
        months = dummy_account.get_spreadsheet_months()
        assert "janvier" in months
        assert "decembre" in months
        assert len(months) == 11

    def test_get_spreadsheet_months_multiuser(self, multiuser_account):
        """Test getting months from multiuser spreadsheet."""
        months = multiuser_account.get_spreadsheet_months()
        assert "janvier" in months
        assert "decembre" in months
        assert len(months) == 11

    def test_get_spreadsheet_categories(self, dummy_account):
        """Test getting categories from spreadsheet."""
        categories = dummy_account.get_spreadsheet_categories()
        assert "foo" in categories
        assert "bar" in categories

    def test_get_spreadsheet_users_monouser(self, dummy_account):
        """Test that monouser spreadsheet returns empty user list."""
        users = dummy_account.get_spreadsheet_users()
        assert users == []

    def test_get_spreadsheet_users_multiuser(self, multiuser_account):
        """Test getting users from multiuser spreadsheet."""
        users = multiuser_account.get_spreadsheet_users()
        assert "alice" in users
        assert "bob" in users
        assert "shared" in users

    def test_is_multiuser_monouser(self, dummy_account):
        """Test that monouser spreadsheet is correctly identified."""
        assert dummy_account.is_multiuser() is False

    def test_is_multiuser_multiuser(self, multiuser_account):
        """Test that multiuser spreadsheet is correctly identified."""
        assert multiuser_account.is_multiuser() is True
